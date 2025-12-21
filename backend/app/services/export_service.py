"""
Export Service

Service for exporting reports in multiple formats including PDF, CSV, Excel, and JSON.
Handles report formatting, file generation, and email delivery.
"""

from typing import Dict, Any, Optional, List, Union, BinaryIO
from datetime import datetime
from io import BytesIO, StringIO
import json
import csv
from pathlib import Path
import tempfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging

# PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.core.config import settings
from app.services.report_service import ReportType, ExportFormat

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting reports in multiple formats"""
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "shikkhaSathi_reports"
        self.temp_dir.mkdir(exist_ok=True)
    
    async def export_report(
        self,
        report_data: Dict[str, Any],
        export_format: ExportFormat,
        filename: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Export report in specified format
        
        Args:
            report_data: Report data to export
            export_format: Format to export to
            filename: Optional custom filename
            
        Returns:
            Export result with file path and metadata
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_type = report_data.get('report_type', 'report')
            filename = f"{report_type}_{timestamp}"
        
        try:
            if export_format == ExportFormat.PDF:
                return await self._export_to_pdf(report_data, filename)
            elif export_format == ExportFormat.CSV:
                return await self._export_to_csv(report_data, filename)
            elif export_format == ExportFormat.EXCEL:
                return await self._export_to_excel(report_data, filename)
            elif export_format == ExportFormat.JSON:
                return await self._export_to_json(report_data, filename)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")
                
        except Exception as e:
            logger.error(f"Export failed for format {export_format}: {str(e)}")
            raise
    
    async def _export_to_pdf(self, report_data: Dict[str, Any], filename: str) -> Dict[str, Any]:
        """Export report to PDF format"""
        if not PDF_AVAILABLE:
            raise RuntimeError("PDF export not available. Install reportlab: pip install reportlab")
        
        file_path = self.temp_dir / f"{filename}.pdf"
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build PDF content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        report_type = report_data.get('report_type', 'Report')
        title = self._get_report_title(report_type, report_data)
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Report period
        if 'report_period' in report_data:
            period = report_data['report_period']
            period_text = f"Report Period: {period['start_date'].strftime('%Y-%m-%d')} to {period['end_date'].strftime('%Y-%m-%d')}"
            story.append(Paragraph(period_text, styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Add content based on report type
        if report_type == ReportType.INDIVIDUAL_STUDENT:
            story.extend(self._build_individual_student_pdf_content(report_data, styles))
        elif report_type == ReportType.CLASS_SUMMARY:
            story.extend(self._build_class_summary_pdf_content(report_data, styles))
        elif report_type == ReportType.COMPARATIVE_ANALYSIS:
            story.extend(self._build_comparative_analysis_pdf_content(report_data, styles))
        
        # Build PDF
        doc.build(story)
        
        return {
            'success': True,
            'file_path': str(file_path),
            'filename': f"{filename}.pdf",
            'format': ExportFormat.PDF,
            'size_bytes': file_path.stat().st_size,
            'generated_at': datetime.utcnow()
        }
    
    async def _export_to_csv(self, report_data: Dict[str, Any], filename: str) -> Dict[str, Any]:
        """Export report to CSV format"""
        file_path = self.temp_dir / f"{filename}.csv"
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header information
            report_type = report_data.get('report_type', 'Report')
            writer.writerow([f"Report Type: {report_type}"])
            
            if 'report_period' in report_data:
                period = report_data['report_period']
                writer.writerow([f"Period: {period['start_date']} to {period['end_date']}"])
            
            writer.writerow([f"Generated: {datetime.utcnow()}"])
            writer.writerow([])  # Empty row
            
            # Write content based on report type
            if report_type == ReportType.INDIVIDUAL_STUDENT:
                self._write_individual_student_csv_content(writer, report_data)
            elif report_type == ReportType.CLASS_SUMMARY:
                self._write_class_summary_csv_content(writer, report_data)
            elif report_type == ReportType.COMPARATIVE_ANALYSIS:
                self._write_comparative_analysis_csv_content(writer, report_data)
        
        return {
            'success': True,
            'file_path': str(file_path),
            'filename': f"{filename}.csv",
            'format': ExportFormat.CSV,
            'size_bytes': file_path.stat().st_size,
            'generated_at': datetime.utcnow()
        }
    
    async def _export_to_excel(self, report_data: Dict[str, Any], filename: str) -> Dict[str, Any]:
        """Export report to Excel format"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("Excel export not available. Install openpyxl: pip install openpyxl")
        
        file_path = self.temp_dir / f"{filename}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        
        # Header information
        row = 1
        report_type = report_data.get('report_type', 'Report')
        ws[f'A{row}'] = f"Report Type: {report_type}"
        ws[f'A{row}'].font = header_font
        row += 1
        
        if 'report_period' in report_data:
            period = report_data['report_period']
            ws[f'A{row}'] = f"Period: {period['start_date']} to {period['end_date']}"
            ws[f'A{row}'].font = normal_font
            row += 1
        
        ws[f'A{row}'] = f"Generated: {datetime.utcnow()}"
        ws[f'A{row}'].font = normal_font
        row += 2
        
        # Write content based on report type
        if report_type == ReportType.INDIVIDUAL_STUDENT:
            row = self._write_individual_student_excel_content(ws, report_data, row, subheader_font, normal_font)
        elif report_type == ReportType.CLASS_SUMMARY:
            row = self._write_class_summary_excel_content(ws, report_data, row, subheader_font, normal_font)
        elif report_type == ReportType.COMPARATIVE_ANALYSIS:
            row = self._write_comparative_analysis_excel_content(ws, report_data, row, subheader_font, normal_font)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': str(file_path),
            'filename': f"{filename}.xlsx",
            'format': ExportFormat.EXCEL,
            'size_bytes': file_path.stat().st_size,
            'generated_at': datetime.utcnow()
        }
    
    async def _export_to_json(self, report_data: Dict[str, Any], filename: str) -> Dict[str, Any]:
        """Export report to JSON format"""
        file_path = self.temp_dir / f"{filename}.json"
        
        # Convert datetime objects to strings for JSON serialization
        json_data = self._serialize_for_json(report_data)
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
        
        return {
            'success': True,
            'file_path': str(file_path),
            'filename': f"{filename}.json",
            'format': ExportFormat.JSON,
            'size_bytes': file_path.stat().st_size,
            'generated_at': datetime.utcnow()
        }
    
    async def send_report_email(
        self,
        export_result: Dict[str, Any],
        recipient_email: str,
        subject: Optional[str] = None,
        message: Optional[str] = None,
        sender_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Send exported report via email
        
        Args:
            export_result: Result from export_report method
            recipient_email: Email address to send to
            subject: Email subject (optional)
            message: Email message body (optional)
            sender_name: Name of sender (optional)
            
        Returns:
            Email sending result
        """
        if not settings.SMTP_HOST:
            raise RuntimeError("SMTP configuration not available")
        
        try:
            # Create message
            msg = MIMEMultipart()
            msg['From'] = f"{sender_name or 'ShikkhaSathi'} <{settings.SMTP_USER}>"
            msg['To'] = recipient_email
            msg['Subject'] = subject or f"Report - {export_result['filename']}"
            
            # Email body
            body = message or f"""
Dear Recipient,

Please find attached your requested report: {export_result['filename']}

Report Details:
- Format: {export_result['format']}
- Generated: {export_result['generated_at']}
- File Size: {export_result['size_bytes']} bytes

Best regards,
ShikkhaSathi Team
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach file
            file_path = Path(export_result['file_path'])
            if file_path.exists():
                with open(file_path, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {export_result["filename"]}'
                )
                msg.attach(part)
            
            # Send email
            server = smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT)
            if settings.SMTP_TLS:
                server.starttls()
            if settings.SMTP_USER and settings.SMTP_PASSWORD:
                server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            
            server.send_message(msg)
            server.quit()
            
            return {
                'success': True,
                'recipient': recipient_email,
                'subject': msg['Subject'],
                'sent_at': datetime.utcnow()
            }
            
        except Exception as e:
            logger.error(f"Email sending failed: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'recipient': recipient_email
            }
    
    def _get_report_title(self, report_type: str, report_data: Dict[str, Any]) -> str:
        """Generate appropriate title for report"""
        if report_type == ReportType.INDIVIDUAL_STUDENT:
            student_name = report_data.get('student_info', {}).get('name', 'Student')
            return f"Individual Student Report - {student_name}"
        elif report_type == ReportType.CLASS_SUMMARY:
            class_name = report_data.get('class_info', {}).get('name', 'Class')
            return f"Class Summary Report - {class_name}"
        elif report_type == ReportType.COMPARATIVE_ANALYSIS:
            return "Comparative Analysis Report"
        else:
            return "Report"
    
    def _serialize_for_json(self, obj: Any) -> Any:
        """Convert objects to JSON-serializable format"""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: self._serialize_for_json(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._serialize_for_json(item) for item in obj]
        else:
            return obj
    
    # PDF content builders
    def _build_individual_student_pdf_content(self, report_data: Dict[str, Any], styles) -> List:
        """Build PDF content for individual student report"""
        story = []
        
        # Student information
        student_info = report_data.get('student_info', {})
        story.append(Paragraph(f"Student: {student_info.get('name', 'N/A')}", styles['Heading2']))
        story.append(Paragraph(f"Email: {student_info.get('email', 'N/A')}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            story.append(Paragraph("Overview Metrics", styles['Heading2']))
            metrics = report_data['overview_metrics']
            
            metrics_data = [
                ['Metric', 'Value'],
                ['Current Level', str(metrics.get('current_level', 'N/A'))],
                ['Total XP', str(metrics.get('total_xp', 'N/A'))],
                ['XP Gained (Period)', str(metrics.get('xp_gained_period', 'N/A'))],
                ['Current Streak', str(metrics.get('current_streak', 'N/A'))],
                ['Time Spent (minutes)', str(metrics.get('total_time_spent', 0) // 60)],
                ['Topics Completed', str(metrics.get('topics_completed', 'N/A'))]
            ]
            
            metrics_table = Table(metrics_data)
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(metrics_table)
            story.append(Spacer(1, 12))
        
        # Performance metrics
        if 'performance_metrics' in report_data:
            story.append(Paragraph("Performance Metrics", styles['Heading2']))
            perf = report_data['performance_metrics']
            
            perf_data = [
                ['Metric', 'Score'],
                ['Average Quiz Score', f"{perf.get('average_quiz_score', 0):.1f}%"],
                ['Average Assessment Score', f"{perf.get('average_assessment_score', 0):.1f}%"],
                ['Overall Average', f"{perf.get('overall_average', 0):.1f}%"]
            ]
            
            perf_table = Table(perf_data)
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(perf_table)
            story.append(Spacer(1, 12))
        
        # Subject performance
        if 'subject_performance' in report_data and report_data['subject_performance']:
            story.append(Paragraph("Subject Performance", styles['Heading2']))
            
            subject_data = [['Subject', 'Average Score', 'Time Spent (min)', 'Topics Completed', 'XP Earned']]
            for subject, data in report_data['subject_performance'].items():
                subject_data.append([
                    subject,
                    f"{data.get('average_score', 0):.1f}%",
                    str(data.get('time_spent', 0) // 60),
                    str(data.get('topics_completed', 0)),
                    str(data.get('xp_earned', 0))
                ])
            
            subject_table = Table(subject_data)
            subject_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(subject_table)
            story.append(Spacer(1, 12))
        
        # Recommendations
        if 'recommendations' in report_data and report_data['recommendations']:
            story.append(Paragraph("Recommendations", styles['Heading2']))
            for i, rec in enumerate(report_data['recommendations'], 1):
                story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        return story
    
    def _build_class_summary_pdf_content(self, report_data: Dict[str, Any], styles) -> List:
        """Build PDF content for class summary report"""
        story = []
        
        # Class information
        class_info = report_data.get('class_info', {})
        story.append(Paragraph(f"Class: {class_info.get('name', 'N/A')}", styles['Heading2']))
        story.append(Paragraph(f"Grade: {class_info.get('grade', 'N/A')} | Subject: {class_info.get('subject', 'N/A')}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            story.append(Paragraph("Class Overview", styles['Heading2']))
            metrics = report_data['overview_metrics']
            
            overview_data = [
                ['Metric', 'Value'],
                ['Total Students', str(metrics.get('total_students', 'N/A'))],
                ['Active Students', str(metrics.get('active_students', 'N/A'))],
                ['Engagement Rate', f"{metrics.get('engagement_rate', 0):.1f}%"],
                ['Class Average', f"{metrics.get('class_average', 0):.1f}%"],
                ['Average Quiz Score', f"{metrics.get('average_quiz_score', 0):.1f}%"],
                ['Total Activities', str(metrics.get('total_activities', 'N/A'))]
            ]
            
            overview_table = Table(overview_data)
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(overview_table)
            story.append(Spacer(1, 12))
        
        # Performance distribution
        if 'performance_distribution' in report_data:
            story.append(Paragraph("Performance Distribution", styles['Heading2']))
            dist = report_data['performance_distribution']
            
            dist_data = [
                ['Performance Level', 'Number of Students'],
                ['Excellent (90%+)', str(dist.get('excellent', 0))],
                ['Good (80-89%)', str(dist.get('good', 0))],
                ['Satisfactory (70-79%)', str(dist.get('satisfactory', 0))],
                ['Needs Improvement (<70%)', str(dist.get('needs_improvement', 0))]
            ]
            
            dist_table = Table(dist_data)
            dist_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(dist_table)
            story.append(Spacer(1, 12))
        
        # Top performers
        if 'top_performers' in report_data and report_data['top_performers']:
            story.append(Paragraph("Top Performers", styles['Heading2']))
            
            top_data = [['Rank', 'Student Name', 'Average Score']]
            for i, student in enumerate(report_data['top_performers'], 1):
                top_data.append([
                    str(i),
                    student.get('student_name', 'N/A'),
                    f"{student.get('average_score', 0):.1f}%"
                ])
            
            top_table = Table(top_data)
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(top_table)
            story.append(Spacer(1, 12))
        
        return story
    
    def _build_comparative_analysis_pdf_content(self, report_data: Dict[str, Any], styles) -> List:
        """Build PDF content for comparative analysis report"""
        story = []
        
        # Comparison overview
        story.append(Paragraph("Comparative Analysis Overview", styles['Heading2']))
        story.append(Paragraph(f"Classes Compared: {report_data.get('classes_compared', 0)}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Comparison metrics
        if 'comparison_metrics' in report_data:
            metrics = report_data['comparison_metrics']
            
            comp_data = [
                ['Metric', 'Value'],
                ['Total Students', str(metrics.get('total_students', 'N/A'))],
                ['Average Engagement', f"{metrics.get('average_engagement', 0):.1f}%"],
                ['Average Performance', f"{metrics.get('average_performance', 0):.1f}%"]
            ]
            
            comp_table = Table(comp_data)
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(comp_table)
            story.append(Spacer(1, 12))
        
        # Class comparisons
        if 'class_comparisons' in report_data and report_data['class_comparisons']:
            story.append(Paragraph("Class Performance Comparison", styles['Heading2']))
            
            class_data = [['Rank', 'Class Name', 'Students', 'Engagement', 'Average Score']]
            for i, class_info in enumerate(report_data['class_comparisons'], 1):
                class_data.append([
                    str(i),
                    class_info.get('class_name', 'N/A'),
                    str(class_info.get('student_count', 0)),
                    f"{class_info.get('engagement_rate', 0):.1f}%",
                    f"{class_info.get('class_average', 0):.1f}%"
                ])
            
            class_table = Table(class_data)
            class_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(class_table)
            story.append(Spacer(1, 12))
        
        return story
    
    # CSV content writers
    def _write_individual_student_csv_content(self, writer, report_data: Dict[str, Any]):
        """Write individual student report content to CSV"""
        # Student information
        if 'student_info' in report_data:
            writer.writerow(['Student Information'])
            student = report_data['student_info']
            writer.writerow(['Name', student.get('name', 'N/A')])
            writer.writerow(['Email', student.get('email', 'N/A')])
            writer.writerow([])
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            writer.writerow(['Overview Metrics'])
            writer.writerow(['Metric', 'Value'])
            metrics = report_data['overview_metrics']
            for key, value in metrics.items():
                writer.writerow([key.replace('_', ' ').title(), str(value)])
            writer.writerow([])
        
        # Performance metrics
        if 'performance_metrics' in report_data:
            writer.writerow(['Performance Metrics'])
            writer.writerow(['Metric', 'Score'])
            perf = report_data['performance_metrics']
            for key, value in perf.items():
                writer.writerow([key.replace('_', ' ').title(), f"{value:.1f}%"])
            writer.writerow([])
        
        # Subject performance
        if 'subject_performance' in report_data and report_data['subject_performance']:
            writer.writerow(['Subject Performance'])
            writer.writerow(['Subject', 'Average Score', 'Time Spent (min)', 'Topics Completed', 'XP Earned'])
            for subject, data in report_data['subject_performance'].items():
                writer.writerow([
                    subject,
                    f"{data.get('average_score', 0):.1f}%",
                    str(data.get('time_spent', 0) // 60),
                    str(data.get('topics_completed', 0)),
                    str(data.get('xp_earned', 0))
                ])
            writer.writerow([])
        
        # Recommendations
        if 'recommendations' in report_data and report_data['recommendations']:
            writer.writerow(['Recommendations'])
            for i, rec in enumerate(report_data['recommendations'], 1):
                writer.writerow([f"{i}.", rec])
    
    def _write_class_summary_csv_content(self, writer, report_data: Dict[str, Any]):
        """Write class summary report content to CSV"""
        # Class information
        if 'class_info' in report_data:
            writer.writerow(['Class Information'])
            class_info = report_data['class_info']
            writer.writerow(['Name', class_info.get('name', 'N/A')])
            writer.writerow(['Grade', class_info.get('grade', 'N/A')])
            writer.writerow(['Subject', class_info.get('subject', 'N/A')])
            writer.writerow([])
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            writer.writerow(['Class Overview'])
            writer.writerow(['Metric', 'Value'])
            metrics = report_data['overview_metrics']
            for key, value in metrics.items():
                display_value = f"{value:.1f}%" if 'rate' in key or 'average' in key else str(value)
                writer.writerow([key.replace('_', ' ').title(), display_value])
            writer.writerow([])
        
        # Performance distribution
        if 'performance_distribution' in report_data:
            writer.writerow(['Performance Distribution'])
            writer.writerow(['Performance Level', 'Number of Students'])
            dist = report_data['performance_distribution']
            writer.writerow(['Excellent (90%+)', str(dist.get('excellent', 0))])
            writer.writerow(['Good (80-89%)', str(dist.get('good', 0))])
            writer.writerow(['Satisfactory (70-79%)', str(dist.get('satisfactory', 0))])
            writer.writerow(['Needs Improvement (<70%)', str(dist.get('needs_improvement', 0))])
            writer.writerow([])
        
        # Top performers
        if 'top_performers' in report_data and report_data['top_performers']:
            writer.writerow(['Top Performers'])
            writer.writerow(['Rank', 'Student Name', 'Average Score'])
            for i, student in enumerate(report_data['top_performers'], 1):
                writer.writerow([
                    str(i),
                    student.get('student_name', 'N/A'),
                    f"{student.get('average_score', 0):.1f}%"
                ])
            writer.writerow([])
        
        # At-risk students
        if 'at_risk_students' in report_data and report_data['at_risk_students']:
            writer.writerow(['At-Risk Students'])
            writer.writerow(['Student Name', 'Average Score'])
            for student in report_data['at_risk_students']:
                writer.writerow([
                    student.get('student_name', 'N/A'),
                    f"{student.get('average_score', 0):.1f}%"
                ])
    
    def _write_comparative_analysis_csv_content(self, writer, report_data: Dict[str, Any]):
        """Write comparative analysis report content to CSV"""
        # Overview
        writer.writerow(['Comparative Analysis Overview'])
        writer.writerow(['Classes Compared', str(report_data.get('classes_compared', 0))])
        writer.writerow([])
        
        # Comparison metrics
        if 'comparison_metrics' in report_data:
            writer.writerow(['Comparison Metrics'])
            writer.writerow(['Metric', 'Value'])
            metrics = report_data['comparison_metrics']
            for key, value in metrics.items():
                display_value = f"{value:.1f}%" if 'average' in key else str(value)
                writer.writerow([key.replace('_', ' ').title(), display_value])
            writer.writerow([])
        
        # Class comparisons
        if 'class_comparisons' in report_data and report_data['class_comparisons']:
            writer.writerow(['Class Performance Comparison'])
            writer.writerow(['Rank', 'Class Name', 'Students', 'Engagement Rate', 'Average Score'])
            for i, class_info in enumerate(report_data['class_comparisons'], 1):
                writer.writerow([
                    str(i),
                    class_info.get('class_name', 'N/A'),
                    str(class_info.get('student_count', 0)),
                    f"{class_info.get('engagement_rate', 0):.1f}%",
                    f"{class_info.get('class_average', 0):.1f}%"
                ])
    
    # Excel content writers
    def _write_individual_student_excel_content(self, ws, report_data: Dict[str, Any], start_row: int, subheader_font, normal_font) -> int:
        """Write individual student report content to Excel worksheet"""
        row = start_row
        
        # Student information
        if 'student_info' in report_data:
            ws[f'A{row}'] = "Student Information"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            student = report_data['student_info']
            ws[f'A{row}'] = "Name"
            ws[f'B{row}'] = student.get('name', 'N/A')
            row += 1
            ws[f'A{row}'] = "Email"
            ws[f'B{row}'] = student.get('email', 'N/A')
            row += 2
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            ws[f'A{row}'] = "Overview Metrics"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'].font = subheader_font
            row += 1
            
            metrics = report_data['overview_metrics']
            for key, value in metrics.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
            row += 1
        
        # Subject performance
        if 'subject_performance' in report_data and report_data['subject_performance']:
            ws[f'A{row}'] = "Subject Performance"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            headers = ['Subject', 'Average Score', 'Time Spent (min)', 'Topics Completed', 'XP Earned']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = subheader_font
            row += 1
            
            for subject, data in report_data['subject_performance'].items():
                ws[f'A{row}'] = subject
                ws[f'B{row}'] = f"{data.get('average_score', 0):.1f}%"
                ws[f'C{row}'] = str(data.get('time_spent', 0) // 60)
                ws[f'D{row}'] = str(data.get('topics_completed', 0))
                ws[f'E{row}'] = str(data.get('xp_earned', 0))
                row += 1
            row += 1
        
        return row
    
    def _write_class_summary_excel_content(self, ws, report_data: Dict[str, Any], start_row: int, subheader_font, normal_font) -> int:
        """Write class summary report content to Excel worksheet"""
        row = start_row
        
        # Class information
        if 'class_info' in report_data:
            ws[f'A{row}'] = "Class Information"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            class_info = report_data['class_info']
            ws[f'A{row}'] = "Name"
            ws[f'B{row}'] = class_info.get('name', 'N/A')
            row += 1
            ws[f'A{row}'] = "Grade"
            ws[f'B{row}'] = class_info.get('grade', 'N/A')
            row += 1
            ws[f'A{row}'] = "Subject"
            ws[f'B{row}'] = class_info.get('subject', 'N/A')
            row += 2
        
        # Overview metrics
        if 'overview_metrics' in report_data:
            ws[f'A{row}'] = "Class Overview"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'].font = subheader_font
            row += 1
            
            metrics = report_data['overview_metrics']
            for key, value in metrics.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                display_value = f"{value:.1f}%" if 'rate' in key or 'average' in key else str(value)
                ws[f'B{row}'] = display_value
                row += 1
            row += 1
        
        # Performance distribution
        if 'performance_distribution' in report_data:
            ws[f'A{row}'] = "Performance Distribution"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            ws[f'A{row}'] = "Performance Level"
            ws[f'B{row}'] = "Number of Students"
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'].font = subheader_font
            row += 1
            
            dist = report_data['performance_distribution']
            performance_levels = [
                ('Excellent (90%+)', dist.get('excellent', 0)),
                ('Good (80-89%)', dist.get('good', 0)),
                ('Satisfactory (70-79%)', dist.get('satisfactory', 0)),
                ('Needs Improvement (<70%)', dist.get('needs_improvement', 0))
            ]
            
            for level, count in performance_levels:
                ws[f'A{row}'] = level
                ws[f'B{row}'] = str(count)
                row += 1
            row += 1
        
        return row
    
    def _write_comparative_analysis_excel_content(self, ws, report_data: Dict[str, Any], start_row: int, subheader_font, normal_font) -> int:
        """Write comparative analysis report content to Excel worksheet"""
        row = start_row
        
        # Overview
        ws[f'A{row}'] = "Comparative Analysis Overview"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        ws[f'A{row}'] = "Classes Compared"
        ws[f'B{row}'] = str(report_data.get('classes_compared', 0))
        row += 2
        
        # Comparison metrics
        if 'comparison_metrics' in report_data:
            ws[f'A{row}'] = "Comparison Metrics"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'].font = subheader_font
            row += 1
            
            metrics = report_data['comparison_metrics']
            for key, value in metrics.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                display_value = f"{value:.1f}%" if 'average' in key else str(value)
                ws[f'B{row}'] = display_value
                row += 1
            row += 1
        
        # Class comparisons
        if 'class_comparisons' in report_data and report_data['class_comparisons']:
            ws[f'A{row}'] = "Class Performance Comparison"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            headers = ['Rank', 'Class Name', 'Students', 'Engagement Rate', 'Average Score']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = subheader_font
            row += 1
            
            for i, class_info in enumerate(report_data['class_comparisons'], 1):
                ws[f'A{row}'] = str(i)
                ws[f'B{row}'] = class_info.get('class_name', 'N/A')
                ws[f'C{row}'] = str(class_info.get('student_count', 0))
                ws[f'D{row}'] = f"{class_info.get('engagement_rate', 0):.1f}%"
                ws[f'E{row}'] = f"{class_info.get('class_average', 0):.1f}%"
                row += 1
        
        return row
    
    def cleanup_temp_files(self, older_than_hours: int = 24):
        """Clean up temporary export files older than specified hours"""
        try:
            cutoff_time = datetime.now() - timedelta(hours=older_than_hours)
            
            for file_path in self.temp_dir.glob("*"):
                if file_path.is_file():
                    file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_time < cutoff_time:
                        file_path.unlink()
                        logger.info(f"Cleaned up temporary file: {file_path}")
                        
        except Exception as e:
            logger.error(f"Error cleaning up temporary files: {str(e)}")